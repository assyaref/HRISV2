"""
Script untuk memperbaiki inkonsistensi database HRIS.
Jalankan: python scripts/fix-database.py

Masalah yang diperbaiki:
1. Ferdi: faceRegistered=true tapi faceDescriptor KOSONG → diset ke false
2. Department/Division/Position hilang untuk karyawan baru (dept-5, div-6, pos-8)
3. USERS dengan employeeId tidak valid (emp-02235-3, emp-03233-2, emp-02235-4)
4. Kolom faceDescriptor/faceRegistered kosong di baris lama
"""

import json
import sys
import os
from copy import copy

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import load_workbook, Workbook

SOURCE = os.path.expandvars(
    r"C:\Users\ASUS TUF\Downloads\Database_HRIS_Lengkap (3).xlsx"
)
OUTPUT = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "output", "Database_HRIS_FIXED.xlsx"
)
OUTPUT = os.path.abspath(OUTPUT)

def read_sheet(ws):
    """Baca semua data dari worksheet, return (headers, rows)"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        r = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else ""
            r[h] = val if val is not None else ""
        data.append(r)
    return headers, data

def write_sheet(ws, headers, data):
    """Tulis headers + data ke worksheet, hapus isi lama"""
    # Hapus semua data
    ws.delete_rows(1, ws.max_row)
    # Tulis headers
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    # Tulis data
    for r, row in enumerate(data, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=row.get(h, ""))

def find_row_idx(data, col, value):
    """Cari index baris berdasarkan nilai kolom"""
    for i, row in enumerate(data):
        if str(row.get(col, "")).strip() == str(value).strip():
            return i
    return -1

def main():
    print(f"📂 Membuka: {SOURCE}")
    wb = load_workbook(SOURCE)
    
    fixes = []

    # ========== PERBAIKAN 1: Ferdi faceDescriptor KOSONG ==========
    print("\n🔍 [1/5] Memeriksa faceDescriptor vs faceRegistered...")
    if 'EMPLOYEE' in wb.sheetnames:
        ws_emp = wb['EMPLOYEE']
        headers_emp, emp_data = read_sheet(ws_emp)
        
        for i, emp in enumerate(emp_data):
            face_desc = str(emp.get('faceDescriptor', '')).strip()
            face_reg = str(emp.get('faceRegistered', '')).strip().lower()
            
            # Cek inconsistent: faceRegistered=true tapi descriptor kosong
            if face_reg == 'true' and (len(face_desc) < 3 or face_desc == '[]'):
                emp_data[i]['faceRegistered'] = 'false'
                print(f"   ❌ PERBAIKAN: {emp.get('fullName')} ({emp.get('employeeId')}) - faceDescriptor KOSONG, set faceRegistered=false")
                fixes.append(f"Set faceRegistered=false untuk {emp['fullName']} (descriptor kosong)")
            
            # Cek inconsistent: ada descriptor tapi faceRegistered=false/tidak ada
            elif len(face_desc) > 3 and face_desc != '[]' and face_reg != 'true':
                emp_data[i]['faceRegistered'] = 'true'
                print(f"   ✅ PERBAIKAN: {emp.get('fullName')} ({emp.get('employeeId')}) - ada descriptor, set faceRegistered=true")
                fixes.append(f"Set faceRegistered=true untuk {emp['fullName']} (descriptor ada)")
        
        write_sheet(ws_emp, headers_emp, emp_data)

    # ========== PERBAIKAN 2: Tambah Department/Division/Position yang hilang ==========
    print("\n🔍 [2/5] Memperbaiki Department/Division/Position yang hilang...")
    
    # Cari employee yang referensi dept/div/pos tidak valid
    missing_depts = set()
    missing_divs = set()
    missing_positions = set()
    
    if 'EMPLOYEE' in wb.sheetnames:
        _, emp_data = read_sheet(wb['EMPLOYEE'])
        
        # Baca existing dept/div/pos
        existing_depts = set()
        if 'DEPARTMENT' in wb.sheetnames:
            _, dept_data = read_sheet(wb['DEPARTMENT'])
            existing_depts = {str(d['id']).strip() for d in dept_data}
        
        existing_divs = set()
        if 'DIVISION' in wb.sheetnames:
            _, div_data = read_sheet(wb['DIVISION'])
            existing_divs = {str(d['id']).strip() for d in div_data}
        
        existing_positions = set()
        if 'POSITION' in wb.sheetnames:
            _, pos_data = read_sheet(wb['POSITION'])
            existing_positions = {str(p['id']).strip() for p in pos_data}
        
        for emp in emp_data:
            dept_id = str(emp.get('departmentId', '')).strip()
            div_id = str(emp.get('divisionId', '')).strip()
            pos_id = str(emp.get('positionId', '')).strip()
            
            if dept_id and dept_id not in existing_depts:
                missing_depts.add(dept_id)
            if div_id and div_id not in existing_divs:
                missing_divs.add(div_id)
            if pos_id and pos_id not in existing_positions:
                missing_positions.add(pos_id)
    
    # Tambah department yang hilang
    if missing_depts and 'DEPARTMENT' in wb.sheetnames:
        ws_dept = wb['DEPARTMENT']
        h_dept, d_dept = read_sheet(ws_dept)
        now = "2026-07-23T10:00:00.000Z"
        for dept_id in sorted(missing_depts):
            code = dept_id.replace('dept-', '').upper()
            d_dept.append({
                'id': dept_id,
                'code': code,
                'name': f'Departemen {code}',
                'description': 'Ditambahkan saat perbaikan database',
                'headId': '',
                'isActive': 'true',
                'createdAt': now
            })
            print(f"   ➕ Tambah Department: {dept_id}")
            fixes.append(f"Tambah Department {dept_id}")
        write_sheet(ws_dept, h_dept, d_dept)
    
    # Tambah division yang hilang
    if missing_divs and 'DIVISION' in wb.sheetnames:
        ws_div = wb['DIVISION']
        h_div, d_div = read_sheet(ws_div)
        now = "2026-07-23T10:00:00.000Z"
        for div_id in sorted(missing_divs):
            code = div_id.replace('div-', '').upper()
            # Cari departmentId dari employee yang pakai division ini
            parent_dept = 'dept-operasional'
            for emp in emp_data:
                if str(emp.get('divisionId', '')).strip() == div_id:
                    parent_dept = str(emp.get('departmentId', 'dept-operasional')).strip()
                    break
            d_div.append({
                'id': div_id,
                'code': code,
                'name': f'Divisi {code}',
                'departmentId': parent_dept,
                'description': 'Ditambahkan saat perbaikan database',
                'isActive': 'true',
                'createdAt': now
            })
            print(f"   ➕ Tambah Division: {div_id}")
            fixes.append(f"Tambah Division {div_id}")
        write_sheet(ws_div, h_div, d_div)
    
    # Tambah position yang hilang
    if missing_positions and 'POSITION' in wb.sheetnames:
        ws_pos = wb['POSITION']
        h_pos, d_pos = read_sheet(ws_pos)
        now = "2026-07-23T10:00:00.000Z"
        for pos_id in sorted(missing_positions):
            code = pos_id.replace('pos-', '').upper()
            # Cari departmentId dari employee yang pakai position ini
            parent_dept = 'dept-operasional'
            for emp in emp_data:
                if str(emp.get('positionId', '')).strip() == pos_id:
                    parent_dept = str(emp.get('departmentId', 'dept-operasional')).strip()
                    break
            d_pos.append({
                'id': pos_id,
                'code': code,
                'name': f'Jabatan {code}',
                'departmentId': parent_dept,
                'level': '1',
                'description': 'Ditambahkan saat perbaikan database',
                'isActive': 'true',
                'createdAt': now
            })
            print(f"   ➕ Tambah Position: {pos_id}")
            fixes.append(f"Tambah Position {pos_id}")
        write_sheet(ws_pos, h_pos, d_pos)

    # ========== PERBAIKAN 3: Fix USERS dengan employeeId tidak valid ==========
    print("\n🔍 [3/5] Memperbaiki USERS dengan employeeId tidak valid...")
    if 'USERS' in wb.sheetnames and 'EMPLOYEE' in wb.sheetnames:
        ws_usr = wb['USERS']
        h_usr, usr_data = read_sheet(ws_usr)
        
        _, emp_data = read_sheet(wb['EMPLOYEE'])
        valid_emp_ids = {str(e['id']).strip() for e in emp_data}
        
        for i, user in enumerate(usr_data):
            emp_id = str(user.get('employeeId', '')).strip()
            if emp_id and emp_id not in valid_emp_ids:
                print(f"   ⚠️  User {user.get('email')} references non-existent employeeId: {emp_id}")
                # Cari employee yang cocok berdasarkan nama
                user_name = str(user.get('name', '')).strip().lower()
                matched = None
                for emp in emp_data:
                    if str(emp.get('fullName', '')).strip().lower() == user_name:
                        matched = emp
                        break
                
                if matched:
                    usr_data[i]['employeeId'] = matched['id']
                    print(f"   ✅ FIX: User {user['email']} employeeId diubah ke {matched['id']} (matched by name: {matched['fullName']})")
                    fixes.append(f"Fix USERS {user['email']}: employeeId {emp_id} -> {matched['id']}")
                else:
                    # Jika tidak ada match, kosongkan employeeId
                    # Tapi untuk admin/HR/Manager, ini normal
                    role = str(user.get('role', '')).strip()
                    if role in ['Administrator', 'HR', 'Manager']:
                        print(f"   ℹ️  Dibiarkan (role={role}, tidak wajib punya employeeId)")
                    else:
                        usr_data[i]['employeeId'] = ''
                        print(f"   ⚠️  EmployeeId dikosongkan (tidak ditemukan match)")
                        fixes.append(f"Kosongkan employeeId {emp_id} untuk user {user['email']}")
        
        write_sheet(ws_usr, h_usr, usr_data)

    # ========== PERBAIKAN 4: Tambahkan kolom faceDescriptor/faceRegistered di EMPLOYEE jika belum ada ==========
    print("\n🔍 [4/5] Memastikan kolom faceDescriptor & faceRegistered ada...")
    if 'EMPLOYEE' in wb.sheetnames:
        ws_emp = wb['EMPLOYEE']
        headers_emp, emp_data = read_sheet(ws_emp)
        
        added = []
        if 'faceDescriptor' not in headers_emp:
            headers_emp.append('faceDescriptor')
            for row in emp_data:
                row['faceDescriptor'] = ''
            added.append('faceDescriptor')
        if 'faceRegistered' not in headers_emp:
            headers_emp.append('faceRegistered')
            for row in emp_data:
                row['faceRegistered'] = 'false'
            added.append('faceRegistered')
        
        if added:
            write_sheet(ws_emp, headers_emp, emp_data)
            print(f"   ➕ Kolom ditambahkan: {', '.join(added)}")
            fixes.append(f"Tambah kolom: {', '.join(added)}")
        else:
            print("   ✅ Kolom sudah lengkap")

    # ========== PERBAIKAN 5: Fix EMPLOYEE yang joinDate kosong atau data tidak lengkap ==========
    print("\n🔍 [5/5] Memperbaiki data EMPLOYEE yang tidak lengkap...")
    if 'EMPLOYEE' in wb.sheetnames:
        ws_emp = wb['EMPLOYEE']
        headers_emp, emp_data = read_sheet(ws_emp)
        
        for i, emp in enumerate(emp_data):
            changed = False
            # joinDate kosong → isi dengan createdAt date
            if not str(emp.get('joinDate', '')).strip():
                created = str(emp.get('createdAt', '')).strip()
                if created:
                    # Ambil bagian tanggal dari ISO timestamp
                    join_date = created[:10]  # YYYY-MM-DD
                else:
                    join_date = '2026-07-17'
                emp_data[i]['joinDate'] = join_date
                changed = True
            
            # employmentStatus kosong → Active
            if not str(emp.get('employmentStatus', '')).strip():
                emp_data[i]['employmentStatus'] = 'Active'
                changed = True
            
            if changed:
                print(f"   ✅ Fix data: {emp.get('fullName')} ({emp.get('employeeId')})")
                fixes.append(f"Fix data employee {emp['fullName']}")
        
        if emp_data:
            write_sheet(ws_emp, headers_emp, emp_data)

    # ========== SIMPAN ==========
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    print(f"\n{'='*60}")
    print(f"✅ DATABASE BERHASIL DIPERBAIKI!")
    print(f"📁 Output: {OUTPUT}")
    print(f"\n📋 Ringkasan Perbaikan ({len(fixes)} items):")
    for f in fixes:
        print(f"   • {f}")
    print(f"\n⚠️  Upload file ini ke Google Spreadsheet Anda!")
    print(f"   Ganti spreadsheet ID: 13gXyJRNeSxyx6pg5fsQISZLljEg8fSJ6nPjKDR2mOYI")
    print(f"   (atau sesuaikan CONFIG.SPREADSHEET_ID di GAS - HRIS/Config.gs)")
    print(f"{'='*60}")

if __name__ == '__main__':
    main()